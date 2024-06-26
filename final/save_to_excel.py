from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import pyodbc
from dotenv import load_dotenv
import os



def trocar_cores(ws, row_index):
    # largura
    for column in ['A', 'D', 'E','I']:
        ws.column_dimensions[column].width = 10
    for column in ['G', 'H']:
        ws.column_dimensions[column].width = 45
    for column in ['A', 'C']:
        ws.column_dimensions[column].width = 6
    for column in ['F','B']:
        ws.column_dimensions[column].width = 18
    for column in ['K','L']:
        ws.column_dimensions[column].width = 16

    ws.column_dimensions['M'].width = 59

    # cores
    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M']:
        fundo = PatternFill(start_color="a6e1fa", end_color="a6e1fa", fill_type="solid")
        ws[f"{column}{row_index}"].fill = fundo
    for column in ['E', 'H']:
        amarelo = PatternFill(start_color="ffea00", end_color="ffea00", fill_type="solid")
        ws[f"{column}{row_index}"].fill = amarelo
    for column in ['C','D', 'G','I','J']:
        marrom = PatternFill(start_color="ffe6a7", end_color="ffe6a7", fill_type="solid")
        ws[f"{column}{row_index}"].fill = marrom


def save_to_excel(nf_pdf_map, nf_zip_map, file_name):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Notas Fiscais'
    ws['A1'] = 'Tipo'
    ws['B1'] = 'Data Email'
    ws['C1'] = 'Série'
    ws['D1'] = 'Nota Fiscal'
    ws['E1'] = 'NF Comp'
    ws['F1'] = 'Data Descarga'
    ws['G1'] = 'Chave de Acesso'
    ws['H1'] = 'Chave NFE Comp'
    ws['I1'] = 'Peso'
    ws['J1'] = 'Lote'
    ws['K1'] = 'Produto'
    ws['L1'] = 'Cnpj Destinatario'
    ws['M1'] = 'Origem'

    row_index = 2
    resultados_sql = conecta_sql(nf_zip_map,nf_pdf_map)

    for nf_formatado, info in nf_pdf_map.items():
        ws.cell(row=row_index, column=1, value="PDF")
        if isinstance(info['data_email'], datetime):
            ws.cell(row=row_index, column=2, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
        else:
            ws.cell(row=row_index, column=2, value=str(info['data_email']))  # Converte para string se não for datetime

        ws.cell(row=row_index, column=3, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
        ws.cell(row=row_index, column=4, value=info.get('nota_fiscal', ''))
        ws.cell(row=row_index, column=5, value=info['nfe'] if info['nfe'] != '0' else '0')
        ws.cell(row=row_index, column=6, value=info['data_emissao'] if info['data_emissao'] != '0' else '0')
        ws.cell(row=row_index, column=7,
                value=info.get('chave_acesso', '0'))  # ['chave_acesso'] if info['chave_acesso'] != '0' else '0')
        ws.cell(row=row_index, column=8, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
        ws.cell(row=row_index, column=9,
                value=info.get('peso_nfe', '0'))
        ws.cell(row=row_index, column=12, value=info['cnpj'] if info['cnpj'] != '0' else '0')
        ws.cell(row=row_index, column=13, value=info['email_vinculado'])
        #trocar_cores(ws, row_index)
        row_index += 1


    #for valor, info in nf_zip_map.items():
    try:
        for (valor, info), resultado in zip(nf_zip_map.items(), resultados_sql):
            data_descarga = str(resultado[8])
            formatoa = '%Y-%m-%d %H:%M:%S.%f'
            formatob = '%d/%m/%Y %H:%M:%S'

            data_formatada = datetime.strptime(data_descarga, formatoa)
            data_ajustada = data_formatada.strftime(formatob)
            ws.cell(row=row_index, column=1, value="ZIP")
            if isinstance(info['data_email'], datetime):
                ws.cell(row=row_index, column=2, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=2, value=str(info['data_email']))
            ws.cell(row=row_index, column=3, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
            ws.cell(row=row_index, column=4, value=info.get('nota_fiscal', 'não encontrei NF'))
            ws.cell(row=row_index, column=5, value=info['nfe'] if info['nfe'] != '0' else '0')
            #ws.cell(row=row_index, column=6, value=info['data_emissao'] if info['data_emissao'] != '0' else '0')

            ws.cell(row=row_index, column=6, value=data_ajustada) #DATA DESCARGA
            ws.cell(row=row_index, column=7, value=info['chave_acesso'] if info['chave_acesso'] != '0' else str(resultado[1]))
            ws.cell(row=row_index, column=8, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
            ws.cell(row=row_index, column=9, value=str(resultado[4]).replace('.',','))  # PESO
            ws.cell(row=row_index, column=10, value=str(resultado[7]))  # LOTE
            ws.cell(row=row_index, column=11, value=str(resultado[9]))  # PRODUTO
            ws.cell(row=row_index, column=12, value=info['cnpj'] if info['cnpj'] != '0' else '0')
            ws.cell(row=row_index, column=13, value=info.get('email_vinculado', ''))

            trocar_cores(ws, row_index)
            row_index += 1

        wb.save(file_name)
        print(f"Planilha '{file_name}' salva com sucesso.")

    except Exception as e:
        print(f"Erro ao salvar a planilha: {e}")


def conecta_sql(nf_zip_map, nf_pdf_map):#, file_name):

    load_dotenv()
    server = os.getenv('DB_SERVER')
    database = os.getenv('DB_DATABASE')
    user = os.getenv('DB_USER')
    senha = os.getenv('DB_PASSWORD')


    consulta_nfe = ', '.join([f"'{info.get('nfe', '0')}'" for valor, info in nf_zip_map.items()]
                                + [f"'{info.get('nfe', '0')}'" for valor, info in nf_pdf_map.items()])
                                #+ [f"'{info.get('nfe', '0')}'" for info in nf_excel_map])

    chaves_consulta = ', '.join([f"'{info.get('nota_fiscal', '0')}'" for valor, info in nf_zip_map.items()]
                                + [f"'{info.get('nota_fiscal', '0')}'" for valor, info in nf_pdf_map.items()])
                                #+ [f"'{info.get('nota_fiscal', '0')}'" for info in nf_excel_map])

    cnpj_consulta = ', '.join([f"'{info.get('cnpj', '0')}'" for valor, info in nf_zip_map.items()]
                                + [f"'{info.get('cnpj', '0')}'" for valor, info in nf_pdf_map.items()])
                                #+ [f"'{info.get('cnpj', '0')}'" for info in nf_excel_map])


    conn_str = f'DRIVER={{SQL SERVER}};SERVER={server};DATABASE={database};UID={user};' \
               f'PWD={senha}'

    #print(chaves_consulta)
    #print(cnpj_consulta)

    try:
        conn = pyodbc.connect(conn_str)
        print('Conectou')
        cursor = conn.cursor()

        # print()
        consulta_sql = f"Select Nota, Chave_nfe, RazaoSocial, cnpj, diferenca, Dt_Emissao, codigo, lote, dtMovimento, Produto" \
                       f" from eis_v_mapa_estoque_python where Nota in ({chaves_consulta}) and cnpj in ({cnpj_consulta}) and FlagCompl = 'N'"
        # consulta_sql = f"Select diferenca from eis_v_mapa_estoque_python where Chave_Nfe on ({chaves_consulta})"
        cursor.execute(consulta_sql)

        resultados = cursor.fetchall()

        cursor_check = conn.cursor()
        consulta_check_sql = f"SELECT Nota FROM eis_v_mapa_estoque_python WHERE Nota IN ({consulta_nfe}) AND cnpj IN ({cnpj_consulta})"
        cursor_check.execute(consulta_check_sql)
        check_resultado = cursor_check.fetchone()
        #print(resultados)

        cursor.close()
        cursor_check.close()
        conn.close()

        return resultados #, check_resultado is not None
        #return resultados

    except pyodbc.Error as e:
        print("erro", e)
        return None,False