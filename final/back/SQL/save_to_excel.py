from PIL.ImImagePlugin import number
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from datetime import datetime
import pyodbc
from dotenv import load_dotenv
import os


def parse_datetime(date_str):
    formats = ['%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Data'{date_str}")


def negativo_cor(ws, row_index):
    #fundo = PatternFill(start_color="FF0101", end_color="FF0101", fill_type="solid")
    fundo = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_index, column=col)
        cell.fill = fundo


def lanc_comp(ws, row_index):
    fundo = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row_index, column=col)
        cell.fill = fundo


def no_desc(ws, row_index):
    if row_index > 1:
        #fundo = PatternFill(start_color="5D736D", end_color="5D736D", fill_type="solid")
        fundo = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
        for col_index in range(5, 19):
            ws.cell(row=row_index, column=col_index, value="NÃO DESCARREGOU")
        for col in range(5, 19):
            cell = ws.cell(row_index, column=col)
            cell.fill = fundo
        #fonte_cor = Font(color="FFFFFF")

def trocar_cores(ws, row_index):
    # largura colunas
    for column in ['A', 'C', 'L', 'F', 'I', 'M','N', 'O', 'Q', 'S']:
        ws.column_dimensions[column].width = 12
    for column in ['K', 'P']:
        ws.column_dimensions[column].width = 45
    for column in ['A', 'J']:
        ws.column_dimensions[column].width = 6
    for column in ['B', 'H', 'G', 'R']:
        ws.column_dimensions[column].width = 18
    for column in ['E']:
        ws.column_dimensions[column].width = 16
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['T'].width = 4
    ws.column_dimensions['U'].width = 28
    ws.column_dimensions['V'].width = 106

    # cores
    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T','U','V']:
        #fundo = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        fundo = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        ws[f"{column}{row_index}"].fill = fundo
    for column in ['H', 'I', 'J', 'K', 'L','T','U']:
        #cor_comp = PatternFill(start_color="C7F9CC", end_color="C7F9CC", fill_type="solid")
        cor_comp = PatternFill(start_color="48D1CC", end_color="48D1CC", fill_type="solid")
        ws[f"{column}{row_index}"].fill = cor_comp
    for column in ['E', 'F', 'G', 'H', 'L', 'O', 'P', 'Q']:
        #cor_nf = PatternFill(start_color="FCBC5D", end_color="FCBC5D", fill_type="solid")
        cor_nf = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        ws[f"{column}{row_index}"].fill = cor_nf


def save_to_excel(nf_pdf_map, nf_zip_map, nf_excel_map, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Notas Fiscais'
    headers = ['Tipo', 'Data Email', 'Emitido', 'Transportadora', 'Cnpj Destinatario',
               'Cod Dest', 'Produto', 'Data Descarga', 'Nota Fiscal', 'Serie',
               'Chave de Acesso', 'Peso Nota', 'Deposito', 'Lote',
               'NF Comp', 'Chave NFE Comp', 'Serie Comp', 'Data Emissao Comp', 'Peso Comp',
               'R/F','Observ',
               'Origem']

    estilo_data = NamedStyle(name='datetime_style', number_format='DD/MM/YYYY HH:MM:SS')

    #cor = PatternFill(start_color="f0812c", end_color="f0812c", fill_type="solid")
    cor = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
    fonte_cor = Font(color="FFFFFF")

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = cor
        cell.font = fonte_cor

    row_index = 2
    (resultados_pdf, resultados_zip, resultados_excel,
     resultados_zip_comp, resultados_pdf_comp, resultados_excel_comp
     ) = conecta_sql(nf_zip_map, nf_pdf_map, nf_excel_map)

    try:
        row_index = 2
        resultados_dict = {str(r[0]): r for r in resultados_pdf}
        lancado_dict = {str(s[0]): s for s in resultados_pdf_comp}
        for nf_formatado, info in nf_pdf_map.items():
            resultado_comp = lancado_dict.get(info['nfe'], None)
            resultado = resultados_dict.get(nf_formatado.lstrip('0'), None)
            if resultado:
                datadesc = resultado[8].strftime('%d/%m/%Y %H:%M:%S')
                dataajusta = datetime.strptime(datadesc, '%d/%m/%Y %H:%M:%S')
            ws.cell(row=row_index, column=1, value="PDF")
            if isinstance(info['data_email'], datetime):
                ws.cell(row=row_index, column=2, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=2,
                        value=str(info['data_email']))
            ws.cell(row=row_index, column=4, value=info.get('transportadora', ''))
            ws.cell(row=row_index, column=5, value=info['cnpj'] if info['cnpj'] != '0' else '0')
            ws.cell(row=row_index, column=6, value=str(resultado[6]) if resultado else 'SEM LEITURA')  # CODDEST
            ws.cell(row=row_index, column=7, value=str(resultado[9]) if resultado else 'SEM LEITURA')  # PRODUTONOTA
            ws.cell(row=row_index, column=8,
                    value=dataajusta if resultado else 'SEM LEITURA').style = estilo_data  # DATADESCARGA
            ws.cell(row=row_index, column=9, value=nf_formatado)
            ws.cell(row=row_index, column=10, value=str(resultado[11]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=11, value=str(resultado[1]) if resultado else 'SEM LEITURA')  # CHAVEACESSO
            ws.cell(row=row_index, column=12, value=resultado[4] if resultado else 'SEM LEITURA')  # PESONOTA
            # value=str(resultado[4]).replace('.', ',')
            ws.cell(row=row_index, column=13, value=str(resultado[14]) if resultado else 'SEM LEITURA') #DEPOSITO
            ws.cell(row=row_index, column=14, value=str(resultado[7]) if resultado else 'SEM LEITURA')  # LOTENOTA
            ws.cell(row=row_index, column=15, value=info['nfe'] if info['nfe'] != '0' else 'SEM LEITURA')
            ws.cell(row=row_index, column=16, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
            ws.cell(row=row_index, column=17, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
            if isinstance(info['data_emissao'], datetime):
                ws.cell(row=row_index, column=18, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=18,
                        value=str(info['data_emissao']))
            ws.cell(row=row_index, column=19, value=info.get('peso_comp', ''))
            ws.cell(row=row_index, column=20, value=str(resultado[12]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=21, value=str(resultado[13]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=22, value=info.get('email_vinculado', ''))
            trocar_cores(ws, row_index)

            if resultado:
                if resultado[4] <= 0:
                    for column in ['L']:
                        #cor_comp = PatternFill(start_color="C40000", end_color="C40000", fill_type="solid")
                        cor_comp = PatternFill(start_color="FF7F50", end_color= "FF7F50", fill_type="solid")
                        ws[f"{column}{row_index}"].fill = cor_comp
            if resultado_comp:
                for column in ['C']:
                    #cor_comp = PatternFill(start_color="10921C", end_color="10921C", fill_type="solid")
                    cor_comp = PatternFill(start_color="00FF7F", end_color="00FF7F", fill_type="solid")
                    ws[f"{column}{row_index}"].fill = cor_comp
                    ws.cell(row=row_index, column=3, value='SIM')
            else:
                for column in ['C']:
                    #cor_comp = PatternFill(start_color="C40000", end_color="C40000", fill_type="solid")
                    cor_comp = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
                    ws[f"{column}{row_index}"].fill = cor_comp
                    ws.cell(row=row_index, column=3, value='NÃO')

            if resultado is None and nf_formatado is not None:
                if ".pdf" in nf_formatado.lower():
                    pass
                else:
                    no_desc(ws, row_index)
                    ws.cell(row=row_index, column=9, value=nf_formatado)
                    ws.cell(row=row_index, column=15, value=info['nfe'] if info['nfe'] != '0' else 'SEM LEITURA')
                    ws.cell(row=row_index, column=16, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
                    ws.cell(row=row_index, column=17, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
                    if isinstance(info['data_emissao'], datetime):
                        ws.cell(row=row_index, column=18, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        ws.cell(row=row_index, column=18,
                                value=str(info['data_emissao']))
            row_index += 1

    except Exception as e:
        print(f"Erro ao salvar a planilha vindo do PDF {e}")

    try:

        resultados_dict = {str(r[0]): r for r in resultados_zip}
        lancado_dict = {str(s[0]): s for s in resultados_zip_comp}
        for valor, info in nf_zip_map.items():
            resultado = resultados_dict.get(valor.lstrip('0'), None)
            resultado_comp = lancado_dict.get(info['nfe'], None)
            if resultado:
                datadesc = resultado[8].strftime('%d/%m/%Y %H:%M:%S')
                datajust = datetime.strptime(datadesc, '%d/%m/%Y %H:%M:%S')
            ws.cell(row=row_index, column=1, value="ZIP")
            if isinstance(info['data_email'], datetime):
                ws.cell(row=row_index, column=2, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=2,
                        value=str(info['data_email']))
            ws.cell(row=row_index, column=4, value=info.get('transportadora', ''))
            ws.cell(row=row_index, column=5, value=info['cnpj'] if info['cnpj'] != '0' else '0')
            ws.cell(row=row_index, column=6, value=str(resultado[6]) if resultado else 'SEM LEITURA')  # CODDEST
            ws.cell(row=row_index, column=7, value=str(resultado[9]) if resultado else 'SEM LEITURA')  # PRODUTONOTA
            ws.cell(row=row_index, column=8,
                    value=datajust if resultado else 'SEM LEITURA').style = estilo_data  # DATADESCARGA
            ws.cell(row=row_index, column=9, value=valor)
            ws.cell(row=row_index, column=10, value=str(resultado[11]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=11, value=str(resultado[1]) if resultado else 'SEM LEITURA')  # CHAVEACESSO
            ws.cell(row=row_index, column=12, value=resultado[4] if resultado else 'SEM LEITURA')  # PESONOTA
            # value=str(resultado[4]).replace('.', ',')

            ws.cell(row=row_index, column=13, value=str(resultado[14]) if resultado else 'SEM LEITURA') #DEPOSITO
            ws.cell(row=row_index, column=14, value=str(resultado[7]) if resultado else 'SEM LEITURA')  # LOTENOTA
            ws.cell(row=row_index, column=15, value=info['nfe'] if info['nfe'] != '0' else 'SEM LEITURA')
            ws.cell(row=row_index, column=16, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
            ws.cell(row=row_index, column=17, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
            if isinstance(info['data_emissao'], datetime):
                ws.cell(row=row_index, column=18, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=18,
                        value=str(info['data_emissao']))
            ws.cell(row=row_index, column=19, value=info.get('peso_comp', ''))
            ws.cell(row=row_index, column=20, value=str(resultado[12]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=21, value=str(resultado[13]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=22, value=info.get('email_vinculado', ''))
            trocar_cores(ws, row_index)
            if resultado:
                if resultado[4] <= 0:
                    for column in ['L']:
                        # cor_comp = PatternFill(start_color="C40000", end_color="C40000", fill_type="solid")
                        cor_comp = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
                        ws[f"{column}{row_index}"].fill = cor_comp
            if resultado_comp:
                for column in ['C']:
                    # cor_comp = PatternFill(start_color="10921C", end_color="10921C", fill_type="solid")
                    cor_comp = PatternFill(start_color="00FF7F", end_color="00FF7F", fill_type="solid")
                    ws[f"{column}{row_index}"].fill = cor_comp
                    ws.cell(row=row_index, column=3, value='SIM')
            else:
                for column in ['C']:
                    # cor_comp = PatternFill(start_color="C40000", end_color="C40000", fill_type="solid")
                    cor_comp = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
                    ws[f"{column}{row_index}"].fill = cor_comp
                    ws.cell(row=row_index, column=3, value='NÃO')
            if resultado is None and valor is not None:
                if ".pdf" in valor.lower():
                    pass
                else:
                    no_desc(ws, row_index)
                    ws.cell(row=row_index, column=9, value=valor)
                    ws.cell(row=row_index, column=15, value=info['nfe'] if info['nfe'] != '0' else 'SEM LEITURA')
                    ws.cell(row=row_index, column=16, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
                    ws.cell(row=row_index, column=17, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
                    if isinstance(info['data_emissao'], datetime):
                        ws.cell(row=row_index, column=18, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        ws.cell(row=row_index, column=18,
                                value=str(info['data_emissao']))
            row_index += 1
            #    lanc_comp(ws,row_index)
    except Exception as e:
        print(f"Erro ao salvar a planilha vinda do ZIP {e}")

    try:
        resultados_dict = {str(r[0]): r for r in resultados_excel}
        lancado_dict = {str(s[0]): s for s in resultados_excel_comp}
        for valor, info in nf_excel_map.items():
            resultado = resultados_dict.get(valor.lstrip('0'),None)
            resultado_comp = lancado_dict.get(info['nfe'], None)
            if resultado:
                datadesc = resultado[8].strftime('%d/%m/%Y %H:%M:%S')
                dataajustd = datetime.strptime(datadesc, '%d/%m/%Y %H:%M:%S')
            ws.cell(row=row_index, column=1, value="EXCEL")
            if isinstance(info['data_email'], datetime):
                ws.cell(row=row_index, column=2, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=2, value=str(info['data_email']))
            ws.cell(row=row_index, column=4, value=info.get('transportadora', ''))
            ws.cell(row=row_index, column=5, value=info['cnpj'] if info['cnpj'] != '0' else '0')
            ws.cell(row=row_index, column=6, value=str(resultado[6]) if resultado else 'SEM LEITURA')  # CODDEST
            ws.cell(row=row_index, column=7, value=str(resultado[9]) if resultado else 'SEM LEITURA')  # PRODUTONOTA
            ws.cell(row=row_index, column=8,
                    value=dataajustd if resultado else 'SEM LEITURA').style = estilo_data  # DATADESCARGA
            ws.cell(row=row_index, column=9, value=valor)
            ws.cell(row=row_index, column=10, value=str(resultado[11]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=11, value=str(resultado[1]) if resultado else 'SEM LEITURA')  # CHAVEACESSO
            ws.cell(row=row_index, column=12, value=resultado[4] if resultado else 'SEM LEITURA')  # PESONOTA
            # value=str(resultado[4]).replace('.', ',')
            ws.cell(row=row_index, column=13, value=str(resultado[14]) if resultado else 'SEM LEITURA') #DEPOSITO
            ws.cell(row=row_index, column=14, value=str(resultado[7]) if resultado else 'SEM LEITURA')  # LOTENOTA
            ws.cell(row=row_index, column=15, value=info['nfe'] if info['nfe'] != '0' else 'SEM LEITURA')
            ws.cell(row=row_index, column=16, value=info['chave_comp'] if info['chave_comp'] != '0' else '0')
            ws.cell(row=row_index, column=17, value=info['serie_nf'] if info['serie_nf'] != '0' else '0')
            if isinstance(info['data_emissao'], datetime):
                ws.cell(row=row_index, column=18, value=info['data_email'].strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_index, column=18,
                        value=str(info['data_emissao']))
            ws.cell(row=row_index, column=19, value=info.get('peso_comp', ''))
            ws.cell(row=row_index, column=20, value=str(resultado[12]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=21, value=str(resultado[13]) if resultado else 'SEM LEITURA')
            ws.cell(row=row_index, column=22, value=info.get('email_vinculado', ''))
            trocar_cores(ws, row_index)
            if resultado:
                if resultado[4] <= 0:
                    for column in ['L']:
                        # cor_comp = PatternFill(start_color="C40000", end_color="C40000", fill_type="solid")
                        cor_comp = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
                        ws[f"{column}{row_index}"].fill = cor_comp
            if resultado_comp:
                for column in ['C']:
                    # cor_comp = PatternFill(start_color="10921C", end_color="10921C", fill_type="solid")
                    cor_comp = PatternFill(start_color="00FF7F", end_color="00FF7F", fill_type="solid")
                    ws[f"{column}{row_index}"].fill = cor_comp
                    ws.cell(row=row_index, column=3, value='SIM')
            else:
                for column in ['C']:
                    # cor_comp = PatternFill(start_color="C40000", end_color="C40000", fill_type="solid")
                    cor_comp = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
                    ws[f"{column}{row_index}"].fill = cor_comp
                    ws.cell(row=row_index, column=3, value='NÃO')
            #if resultado is None and valor is not None:
            #    no_desc(ws, row_index)
            #    ws.cell(row=row_index, column=9, value=valor)
            row_index += 1

    except Exception as e:
        print(f"Erro ao salvar a planilha vinda do EXCEL {e}")

    wb.save(file_name)
    print(f"Planilha '{file_name}' salva com sucesso.")
    return len(nf_pdf_map.items()), len(resultados_pdf), len(nf_zip_map.items()), len(resultados_zip), \
        len(nf_excel_map.items()), len(resultados_excel)


def conecta_sql(nf_zip_map, nf_pdf_map, nf_excel_map):  # , file_name):

    load_dotenv()
    server = os.getenv('DB_SERVER')
    database = os.getenv('DB_DATABASE')
    user = os.getenv('DB_USER')
    senha = os.getenv('DB_PASSWORD')

    chaves_consulta_zip = [f"'{info.get('nota_fiscal', '0')}'" for valor, info in nf_zip_map.items()]
    chaves_consulta_pdf = [f"'{info.get('nota_fiscal', '0')}'" for valor, info in nf_pdf_map.items()]
    chaves_consulta_excel = [f"'{info.get('nota_fiscal', '0')}'" for valor, info in nf_excel_map.items()]

    cnpj_consulta_zip = ', '.join([f"'{info.get('cnpj', '0')}'" for valor, info in nf_zip_map.items()])
    cnpj_consulta_pdf = ', '.join([f"'{info.get('cnpj', '0')}'" for valor, info in nf_pdf_map.items()])
    cnpj_consulta_excel = ', '.join([f"'{info.get('cnpj', '0')}'" for valor, info in nf_excel_map.items()])

    emitente_consulta_zip = ', '.join([f"'{info.get('emitente', '0')}'" for valor, info in nf_zip_map.items()])

    # consulta para verificar se já foi lançado a nota
    comp_consulta_pdf = ', '.join([f"'{info.get('chave_comp', '0')}'" for valor, info in nf_pdf_map.items()])
    nfe_consulta_pdf = [f"'{info.get('nfe', '0')}'" for valor, info in nf_pdf_map.items()]

    comp_consulta_zip = ', '.join([f"'{info.get('chave_comp', '0')}'" for valor, info in nf_zip_map.items()])
    nfe_consulta_zip = [f"'{info.get('nfe', '0')}'" for valor, info in nf_zip_map.items()]

    comp_consulta_excel = ', '.join([f"'{info.get('chave_comp', '0')}'" for valor, info in nf_excel_map.items()])
    nfe_consulta_excel = [f"'{info.get('nfe', '0')}'" for valor, info in nf_excel_map.items()]

    olam_consulta_excel = ' ,'.join([f"'{info.get('chave_acesso','0')}'" for valor, info in nf_excel_map.items()])
    conn_str = f'DRIVER={{SQL SERVER}};SERVER={server};DATABASE={database};UID={user};' \
               f'PWD={senha}'

    try:
        conn = pyodbc.connect(conn_str)
        print('Conectou')
        cursor_pdf = conn.cursor()
        cursor_pdf_comp = conn.cursor()
        cursor_zip = conn.cursor()
        cursor_zip_comp = conn.cursor()
        cursor_excel = conn.cursor()
        cursor_excel_comp = conn.cursor()


        consulta_pdf = f"""
        WITH ChavesNumeradas AS (
            SELECT
                Nota,
                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS Posicao
            FROM (VALUES {', '.join([f"({nota})" for nota in chaves_consulta_pdf])}) AS V(Nota)
        ),
            UltNota AS (SELECT
                m.Nota,
                m.Chave_nfe,
                m.RazaoSocial,
                m.cnpj,
                m.diferenca,
                m.Dt_Emissao,
                m.codigo,
                m.lote,
                m.dtMovimento,
                m.Produto,
                o.Posicao,
                m.SerieNF,
                m.tipotransp,
                m.observ,
                m.deposito,
                ROW_NUMBER() OVER (PARTITION BY m.Nota ORDER BY m.Dt_Emissao DESC) AS rn
            FROM
                eis_v_mapa_estoque_python m
            JOIN
                ChavesNumeradas o ON m.Nota = o.Nota
            WHERE
                m.cnpj IN ({cnpj_consulta_pdf})
                
                and m.TipoNota = 'NOTA_FISCAL'
                and m.diferenca < 10000

            )
        SELECT
            Nota,
            Chave_nfe,
            RazaoSocial,
            cnpj,
            diferenca,
            Dt_Emissao,
            codigo,
            lote,
            dtMovimento,
            Produto,
            Posicao,
            SerieNF,
            tipotransp,
            observ,
            deposito
        From
            UltNota
        where
            rn = 1
        ORDER BY
            Posicao;
        """
        # and m.cnpjemitente in ({emitente_consulta_zip})
        consulta_zip = f"""
                WITH ChavesNumeradas AS (
                    SELECT
                        Nota,
                        ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS Posicao
                    FROM (VALUES {', '.join([f"({nota})" for nota in chaves_consulta_zip])}) AS V(Nota)
                ),
                    UltNota AS (SELECT
                        m.Nota,
                        m.Chave_nfe,
                        m.RazaoSocial,
                        m.cnpj,
                        m.diferenca,
                        m.Dt_Emissao,
                        m.codigo,
                        m.lote,
                        m.dtMovimento,
                        m.Produto,
                        o.Posicao,
                        m.SerieNF,
                        m.tipotransp,
                        m.observ,
                        m.deposito,
                        ROW_NUMBER() OVER (PARTITION BY m.Nota ORDER BY m.Dt_Emissao DESC) AS rn
                    FROM
                        eis_v_mapa_estoque_python m
                    JOIN
                        ChavesNumeradas o ON m.Nota = o.Nota
                    WHERE
                        m.cnpj IN ({cnpj_consulta_zip})
                        and m.diferenca < 10000
                        
                    )
                SELECT
                    Nota,
                    Chave_nfe,
                    RazaoSocial,
                    cnpj,
                    diferenca,
                    Dt_Emissao,
                    codigo,
                    lote,
                    dtMovimento,
                    Produto,
                    Posicao,
                    SerieNF,
                    tipotransp,
                    observ,
                    deposito
                From
                    UltNota
                where
                    rn = 1
                ORDER BY
                    Posicao;
                """
        pdf_comp = f"""
                        WITH ChavesNumeradas AS (
                            SELECT
                                Nota,
                                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS Posicao
                            FROM (VALUES {', '.join([f"({nfe})" for nfe in nfe_consulta_pdf])}) AS V(Nota)
                        ),
                            UltNota AS (SELECT
                                m.Nota,
                                m.Chavenfe,
                                O.Posicao,
                                ROW_NUMBER() OVER (PARTITION BY m.Nota ORDER BY m.DtEmissao DESC) AS rn
                            FROM
                                NotaFiscal m
                            JOIN
                                ChavesNumeradas o ON m.Nota = o.Nota
                            left join Destinatario de on de.CodDest = m.CodDest
                            WHERE
                                de.CGCCPF IN ({cnpj_consulta_pdf})
                                and m.chavenfe in ({comp_consulta_pdf})
                            )
                        SELECT
                            Nota,
                            Chavenfe,
                            Posicao
                        From
                            UltNota
                        where
                            rn = 1
                        ORDER BY
                            Posicao;
                        """
        zip_comp = f"""
                        WITH ChavesNumeradas AS (
                            SELECT
                                Nota,
                                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS Posicao
                            FROM (VALUES {', '.join([f"({nfe})" for nfe in nfe_consulta_zip])}) AS V(Nota)
                        ),
                            UltNota AS (SELECT
                                m.Nota,
                                m.Chavenfe,
                                O.Posicao,
                                ROW_NUMBER() OVER (PARTITION BY m.Nota ORDER BY m.DtEmissao DESC) AS rn
                            FROM
                                NotaFiscal m
                            JOIN
                                ChavesNumeradas o ON m.Nota = o.Nota
                            left join Destinatario de on de.CodDest = m.CodDest
                            WHERE
                                de.CGCCPF IN ({cnpj_consulta_zip})
                                and m.chavenfe in ({comp_consulta_zip})
                            )
                        SELECT
                            Nota,
                            Chavenfe,
                            Posicao
                        From
                            UltNota
                        where
                            rn = 1
                        ORDER BY
                            Posicao;
                        """
        excel_comp = f"""
                               WITH ChavesNumeradas AS (
                                   SELECT
                                       Nota,
                                       ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS Posicao
                                   FROM (VALUES {', '.join([f"({nfe})" for nfe in nfe_consulta_excel])}) AS V(Nota)
                               ),
                                   UltNota AS (SELECT
                                       m.Nota,
                                       m.Chavenfe,
                                       O.Posicao,
                                       ROW_NUMBER() OVER (PARTITION BY m.Nota ORDER BY m.DtEmissao DESC) AS rn
                                   FROM
                                       NotaFiscal m
                                   JOIN
                                       ChavesNumeradas o ON m.Nota = o.Nota
                                   left join Destinatario de on de.CodDest = m.CodDest
                                   WHERE
                                       de.CGCCPF IN ({cnpj_consulta_excel})
                                       and m.chavenfe in ({comp_consulta_excel})
                                   )
                               SELECT
                                   Nota,
                                   Chavenfe,
                                   Posicao
                               From
                                   UltNota
                               where
                                   rn = 1
                               ORDER BY
                                   Posicao;
                               """
        consulta_excel = f"""
                        WITH ChavesNumeradas AS (
                            SELECT
                                Nota,
                                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS Posicao
                            FROM (VALUES {', '.join([f"({nota})" for nota in chaves_consulta_excel])}) AS V(Nota)
                        ),
                            UltNota AS (SELECT
                                m.Nota,
                                m.Chave_nfe,
                                m.RazaoSocial,
                                m.cnpj,
                                m.diferenca,
                                m.Dt_Emissao,
                                m.codigo,
                                m.lote,
                                m.dtMovimento,
                                m.Produto,
                                o.Posicao,
                                m.SerieNF,
                                m.tipotransp,
                                m.observ,
                                m.deposito,
                                ROW_NUMBER() OVER (PARTITION BY m.Nota ORDER BY m.Dt_Emissao DESC) AS rn
                            FROM
                                eis_v_mapa_estoque_python m
                            JOIN
                                ChavesNumeradas o ON m.Nota = o.Nota
                            WHERE
                                m.cnpj IN ({cnpj_consulta_excel})
                                --and m.diferenca < 10000
                                and (CASE
									when exists (select 1 from eis_v_mapa_estoque_python where Chave_nfe in ({olam_consulta_excel}))
									then m.chave_nfe
									end) = m.chave_nfe
                            )
                        SELECT
                            Nota,
                            Chave_nfe,
                            RazaoSocial,
                            cnpj,
                            diferenca,
                            Dt_Emissao,
                            codigo,
                            lote,
                            dtMovimento,
                            Produto,
                            Posicao,
                            SerieNF,
                            tipotransp,
                            observ,
                            deposito
                        From
                            UltNota
                        where
                            rn = 1
                        ORDER BY
                            Posicao;
                        """


        resultado_pdf, resultado_zip, resultado_excel, resultado_zip_comp, resultado_pdf_comp, resultado_excel_comp = [], [], [], [], [], []

        try:
            cursor_pdf.execute(consulta_pdf)
            resultado_pdf = cursor_pdf.fetchall()
            cursor_pdf_comp.execute(pdf_comp)
            resultado_pdf_comp = cursor_pdf_comp.fetchall()
        except Exception as e:
            print(f"Erro ao executar a consulta PDF: {e}")
        try:
            cursor_zip.execute(consulta_zip)
            resultado_zip = cursor_zip.fetchall()
            cursor_zip_comp.execute(zip_comp)
            resultado_zip_comp = cursor_zip_comp.fetchall()
        except Exception as e:
            print(f"Erro ao executar a consulta ZIP: {e}")
        try:
            cursor_excel.execute(consulta_excel)
            resultado_excel = cursor_excel.fetchall()
            cursor_excel_comp.execute(excel_comp)
            resultado_excel_comp = cursor_excel_comp.fetchall()
        except Exception as e:
            print(f"Erro ao executar a consulta EXCEL: {e}")

        # Comentei a cima para testar velocidade, chamando todos os cursores em um só try/except, o retorno é mais rápido
        # porém, se der erro em um dos cursores, vai apontar todos como se estivessem errados
        # ganha em velocidade, perda em validação
        """try:
            cursor_pdf.execute(consulta_pdf)
            resultado_pdf = cursor_pdf.fetchall()
            cursor_pdf_comp.execute(pdf_comp)
            resultado_pdf_comp = cursor_pdf_comp.fetchall()

            cursor_zip.execute(consulta_zip)
            resultado_zip = cursor_zip.fetchall()
            cursor_zip_comp.execute(zip_comp)
            resultado_zip_comp = cursor_zip_comp.fetchall()

            cursor_excel.execute(consulta_excel)
            resultado_excel = cursor_excel.fetchall()
            cursor_excel_comp.execute(excel_comp)
            resultado_excel_comp = cursor_excel_comp.fetchall()
        except Exception as e:
            print(f"Erro ao executar os cursores {e}")"""

        cursor_pdf.close()
        cursor_pdf_comp.close()
        cursor_zip.close()
        cursor_zip_comp.close()
        cursor_excel.close()
        cursor_excel_comp.close()

        conn.close()

        if not resultado_pdf:
            print("Nenhum dado encontrado para PDF.")
        if not resultado_zip:
            print("Nenhum dado encontrado para ZIP.")
        if not resultado_excel:
            print("Nenhum dado encontrado para EXCEL.")

        if not resultado_zip_comp:
            print("Nenhum NFE encontrado para ZIP.")

        if not resultado_pdf_comp:
            print("Nenhum NFE encontrado para PDF")

        if not resultado_excel_comp:
            print("Nenhum NFE encontrado para EXCEL")

        return (resultado_pdf, resultado_zip, resultado_excel, resultado_zip_comp, resultado_pdf_comp,
                resultado_excel_comp)



    except Exception as e:
        print(f"Erro ao preparar ou fechar a conexão: {e}")
        return None, None, None, None, None, None
