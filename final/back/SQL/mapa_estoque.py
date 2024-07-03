from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import pymssql
import time
import os
from dotenv import load_dotenv

load_dotenv()
server = os.getenv('DB_SERVER')
database = os.getenv('DB_DATABASE')
user = os.getenv('DB_USER')
senha = os.getenv('DB_PASSWORD')

def start_mapa(output_dir, data_inicio, data_fim):

    db_config = {
        'server': server,
        'user': user,
        'password': senha,
        'database': database

    }
    connection = None
    #conn_str = f'DRIVER={{SQL SERVER}};SERVER={server};DATABASE={database};UID={user};' \
    #           f'PWD={senha}'

    wb = Workbook()
    ws = wb.active
    tempo_inicial = time.time()
    data_inicio_str = datetime.strptime(data_inicio, "%d/%m/%Y %H:%M:%S")
    data_inicio_formatada = data_inicio_str.strftime("%Y-%m-%d %H:%M:%S")

    data_fim_str = datetime.strptime(data_fim, "%d/%m/%Y %H:%M:%S")
    data_fim_formatada = data_fim_str.strftime("%Y-%m-%d %H:%M:%S")

    print(f'{data_inicio_formatada} {data_fim_formatada}')
    try:
        connection = pymssql.connect(**db_config)
        cursor = connection.cursor()
        cursor.execute(f"exec CliRochaRelExpMapaEstRobo 0,'','{data_inicio_formatada}','{data_fim_formatada}'")
        print(f'Data Inicio {data_inicio}, data Fim {data_fim}')

        resultado = cursor.fetchall()

        colunas = [col[0] for col in cursor.description]
        for col_idx, col_name in enumerate(colunas, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')

        # Preencher os dados
        for row_idx, row_data in enumerate(resultado, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Salvar o workbook
        wb.save(os.path.join(output_dir, "mapadeestoque.xlsx"))
        # wb.save("mapadeestoque.xlsx")
        return True
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
    finally:
        if connection is not None:
            connection.close()
            print("Conexão encerrada")
